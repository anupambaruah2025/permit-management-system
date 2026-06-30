from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_permit_checklists_with_forms():
    """Create detailed checklists for all 23 permits with form links"""
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Define checklist items with form links
    checklists = {
        "1. Company Registration": {
            "Authority": "Department of Business Development (DBD)",
            "FormLink": "https://www.dbd.go.th/th/register-business",
            "Items": [
                ["Completed Company Registration Form", "PDF + Original", "https://www.dbd.go.th/th/register-business"],
                ["Certificate of Incorporation", "Original", "https://www.dbd.go.th"],
                ["Articles of Association", "PDF + Certified Copy", "https://www.dbd.go.th"],
                ["Board Resolution - Business Registration", "Signed + Sealed", "N/A - Internal"],
                ["Proof of Registered Address", "Lease or Title Deed", "N/A - Land Documents"],
                ["National ID/Passport of Directors", "Certified Copy", "N/A - Government ID"],
                ["Company Seal", "Physical + Image", "N/A - Company Document"],
                ["Shareholder Information", "Name, ID, Ownership %", "N/A - Internal"],
                ["Initial Registered Capital", "Bank Statement", "N/A - Bank Document"],
                ["Tax ID Application (Form PP.01)", "Completed Form", "https://www.rd.go.th/"],
            ]
        },
        "2. BOI Promotion Certificate": {
            "Authority": "Thailand Board of Investment (BOI)",
            "FormLink": "https://boi-investment.boi.go.th/public/",
            "Items": [
                ["Company Registration Certificate", "Recent (< 6 months)", "https://www.dbd.go.th"],
                ["Articles of Association", "Certified Copy", "https://www.dbd.go.th"],
                ["Board Resolution - BOI Application", "Signed + Sealed", "N/A - Internal"],
                ["Power of Attorney", "Notarized + Apostilled", "https://www.nla.go.th"],
                ["Shareholder List (Form Bor.Or.Jor. 5)", "DBD Certified", "https://www.dbd.go.th"],
                ["Financial Statements", "3-year history, Audited", "N/A - Auditor"],
                ["Bank Certificate - Capital Deposit", "Original from Bank", "N/A - Bank Document"],
                ["Feasibility Study", "Professional, 20+ pages", "N/A - Consultant"],
                ["Technical Specifications", "MEP Drawings, PUE Analysis", "N/A - Engineering"],
                ["Thailand Benefit Enhancement Plan", "Training, Supply Chain", "https://boi-investment.boi.go.th/public/"],
                ["Environmental Assessment", "Phase 1 or Full EIA", "https://www.onep.go.th"],
                ["Proof of Foreign Investment Capacity", "Bank statements, Tax docs", "N/A - Government/Bank"],
                ["Certificate of No Criminal Record", "Apostilled + Translated", "https://www.police.go.th/"],
                ["Passport Copy - Foreign Investor", "Certified + Apostilled", "https://www.mea.go.th/"],
                ["Tax ID Document - Foreign Investor", "Certified + Apostilled", "https://www.rd.go.th/"],
                ["Project Timeline", "Detailed Gantt chart", "N/A - Internal"],
                ["Site Location Map", "1:50,000 scale", "https://www.dlt.go.th/"],
                ["Organizational Chart", "Management structure", "N/A - Internal"],
            ]
        },
        "3. Foreign Business License": {
            "Authority": "Department of Business Development (DBD)",
            "FormLink": "https://www.dbd.go.th/th/foreign-business-license",
            "Items": [
                ["Company Registration Certificate", "Recent (< 6 months)", "https://www.dbd.go.th"],
                ["Certificate of Good Standing", "Apostilled + Translated", "https://www.dbd.go.th"],
                ["Company Affidavit", "Signed + Sealed", "https://www.dbd.go.th"],
                ["Board Resolution - FBL Application", "Signed + Sealed", "N/A - Internal"],
                ["Shareholder Register Form", "Bor.Or.Jor. 5, Certified", "https://www.dbd.go.th"],
                ["Passport Copy - Foreign Investor", "Certified + Apostilled", "https://www.mea.go.th/"],
                ["Certificate of No Criminal Record", "Apostilled + Translated", "https://www.police.go.th/"],
                ["Foreign Investor Tax Documents", "Tax ID + Recent Filing", "https://www.rd.go.th/"],
                ["Bank Certificate - Capital", "THB 2 million minimum", "N/A - Bank Document"],
                ["Business Plan in Thai", "5-10 pages, Detailed", "N/A - Internal"],
                ["Office/Land Documents", "Lease or Title Deed", "https://www.dlt.go.th/"],
                ["Memorandum of Association", "Certified Copy", "https://www.dbd.go.th"],
                ["List of Foreign Investors", "Names, IDs, Nationalities", "N/A - Internal"],
                ["Financial Capacity Proof", "Bank statements, Audited financials", "N/A - Bank/Auditor"],
            ]
        },
        "4. Land Use & Zoning Confirmation": {
            "Authority": "Department of Public Works / Local Authority",
            "FormLink": "https://www.dlt.go.th/",
            "Items": [
                ["Land Title Deed (Chanote)", "Certified Copy", "https://www.dlt.go.th/"],
                ["OR Draft Title (Nor Sor 3)", "If Chanote unavailable", "https://www.dlt.go.th/"],
                ["Title Search Certificate", "Recent (< 3 months)", "https://www.dlt.go.th/"],
                ["Land Map", "1:4,000 scale, Colored", "https://www.dlt.go.th/"],
                ["Site Plan", "Showing boundaries, dimensions", "N/A - Engineering"],
                ["Zoning Certificate Request", "Application form", "https://www.dlt.go.th/"],
                ["Current Land Use Documentation", "Photos, Description", "N/A - Internal"],
                ["Building Plans Summary", "Brief description", "N/A - Engineering"],
                ["Lease Agreement", "If leasing (30-year min)", "N/A - Legal Document"],
                ["Town Planning Map", "From Local Authority", "https://www.dlt.go.th/"],
                ["Environmental Site Assessment", "Phase 1 Report", "https://www.onep.go.th"],
            ]
        },
        "5. Environmental Impact Assessment (EIA)": {
            "Authority": "Office of Natural Resources & Environmental Policy (ONEP)",
            "FormLink": "https://www.onep.go.th/",
            "Items": [
                ["EIA Screening Request", "Project description", "https://www.onep.go.th/"],
                ["Terms of Reference (TOR)", "Approved by ONEP", "https://www.onep.go.th/"],
                ["Full EIA Report", "80-150 pages, Professional", "https://www.onep.go.th/"],
                ["Baseline Environmental Data", "Air, water, noise, ecology", "https://www.onep.go.th/"],
                ["Impact Assessment", "Construction & operational impacts", "https://www.onep.go.th/"],
                ["Mitigation Measures", "Detailed technical specs", "https://www.onep.go.th/"],
                ["Water Management Plan", "Consumption & discharge", "https://www.onep.go.th/"],
                ["Waste Management Plan", "Hazardous & non-hazardous", "https://www.onep.go.th/"],
                ["Electrical & Generator Emissions", "Stack testing procedure", "https://www.onep.go.th/"],
                ["Noise & Vibration Assessment", "Modeling & measurements", "https://www.onep.go.th/"],
                ["Traffic Impact Study", "Construction & operation", "https://www.onep.go.th/"],
                ["Community Consultation Records", "Meeting minutes, attendance", "https://www.onep.go.th/"],
                ["Alternatives Analysis", "Location, technology options", "https://www.onep.go.th/"],
                ["Monitoring Plan", "5+ year compliance monitoring", "https://www.onep.go.th/"],
                ["EIA Agency Approval Certificate", "ONEP formal approval", "https://www.onep.go.th/"],
            ]
        },
        "6. Building Construction Permit": {
            "Authority": "Local Building Control Authority (AHJ)",
            "FormLink": "https://www.dlt.go.th/",
            "Items": [
                ["Building Construction Permit Application", "Completed form", "https://www.dlt.go.th/"],
                ["Architectural Plans", "Scale 1:100, Floor plans", "N/A - Engineering"],
                ["Structural Engineering Plans", "Calculations, Drawings", "N/A - Engineering"],
                ["Electrical Installation Drawings", "Single-line diagram", "N/A - Engineering"],
                ["MEP (Mechanical/Electrical/Plumbing)", "Detailed specifications", "N/A - Engineering"],
                ["Fire Safety Plans", "Fire suppression, exits, lighting", "N/A - Engineering"],
                ["Building Layout", "All dimensions, setbacks", "N/A - Engineering"],
                ["Environmental Impact Summary", "EIA Approval letter", "https://www.onep.go.th"],
                ["Land Title/Lease Documents", "Proof of land rights", "https://www.dlt.go.th/"],
                ["Site Location Map", "1:50,000 scale", "https://www.dlt.go.th/"],
                ["Proof of Land Ownership/Control", "Title deed or lease", "https://www.dlt.go.th/"],
                ["Building Design by Licensed Architect", "Architect seal & signature", "N/A - Professional"],
                ["Structural Design by Licensed Engineer", "Engineer seal & signature", "N/A - Professional"],
                ["Board Resolution - Construction", "Company authorized", "N/A - Internal"],
                ["Zoning Compliance Certificate", "From Local Authority", "https://www.dlt.go.th/"],
            ]
        },
        "7. Road Access / Highway Connection": {
            "Authority": "Local Authority / Highway Authority",
            "FormLink": "https://www.dlt.go.th/",
            "Items": [
                ["Road Access Application", "Formal request", "https://www.dlt.go.th/"],
                ["Site Location Map", "Showing road access point", "https://www.dlt.go.th/"],
                ["Site Plan", "Access route, dimensions", "N/A - Engineering"],
                ["Engineering Plans", "Road construction details", "N/A - Engineering"],
                ["Environmental Assessment", "Road impact analysis", "https://www.onep.go.th"],
                ["Traffic Impact Study", "Vehicle counts, patterns", "N/A - Consultant"],
                ["Road Safety Plan", "Signage, markings, safety", "N/A - Engineering"],
                ["Proof of Land Rights", "Title deed or lease", "https://www.dlt.go.th/"],
                ["Building Permit Approval", "BCA approval copy", "https://www.dlt.go.th/"],
                ["Engineering Certification", "Licensed engineer approval", "N/A - Professional"],
            ]
        },
        "8. Grid Connection Approval": {
            "Authority": "Provincial Electricity Authority (PEA) / MEA",
            "FormLink": "https://www.pea.co.th/ or https://www.mea.or.th",
            "Items": [
                ["Grid Connection Application", "Completed form", "https://www.pea.co.th/ or https://www.mea.or.th"],
                ["Load Calculation", "Total MW required", "N/A - Engineering"],
                ["Single-Line Electrical Diagram", "Transformer, distribution", "N/A - Engineering"],
                ["Substation Design", "Location, specifications", "N/A - Engineering"],
                ["Power Distribution Design", "Cable routing, sizing", "N/A - Engineering"],
                ["Environmental Assessment", "Electrical infrastructure", "https://www.onep.go.th"],
                ["Site Map", "Showing proposed substation", "https://www.dlt.go.th/"],
                ["Engineering Design", "By licensed electrical engineer", "N/A - Professional"],
                ["Proof of Land Control", "For substation location", "https://www.dlt.go.th/"],
                ["Project Timeline", "Construction & operational dates", "N/A - Internal"],
                ["Coordination with Building Permit", "Building approval", "https://www.dlt.go.th/"],
            ]
        },
        "9. Power Availability Confirmation": {
            "Authority": "Energy Regulatory Commission (ERC)",
            "FormLink": "https://www.erc.or.th/",
            "Items": [
                ["Power Requirement Letter", "Total capacity needed", "https://www.erc.or.th/"],
                ["Load Forecast", "Peak and average loads", "N/A - Engineering"],
                ["Project Specifications", "Type, size, timeline", "N/A - Internal"],
                ["Site Location Map", "Showing facility location", "https://www.dlt.go.th/"],
                ["Grid Connection Plans", "Transformer, substation", "N/A - Engineering"],
                ["Environmental Compliance", "EIA approval", "https://www.onep.go.th"],
                ["Proof of Applicant Identity", "Company registration", "https://www.dbd.go.th"],
                ["Power Usage Projections", "Monthly/annual estimates", "N/A - Internal"],
            ]
        },
        "10. High-Voltage Substation Approval": {
            "Authority": "PEA / MEA",
            "FormLink": "https://www.pea.co.th/ or https://www.mea.or.th",
            "Items": [
                ["Substation Design", "Professional engineering drawings", "https://www.pea.co.th/ or https://www.mea.or.th"],
                ["Electrical Specifications", "Voltage, capacity, equipment", "N/A - Engineering"],
                ["Safety Plan", "Personnel protection, procedures", "N/A - Engineering"],
                ["Site Plan", "Showing substation location", "https://www.dlt.go.th/"],
                ["Proof of Land Control", "Title deed or lease", "https://www.dlt.go.th/"],
                ["Environmental Assessment", "EMF, noise, visual impact", "https://www.onep.go.th"],
                ["Grid Connection Plan", "Connection to main grid", "https://www.pea.co.th/ or https://www.mea.or.th"],
                ["Licensed Engineer Certification", "Design review & approval", "N/A - Professional"],
                ["Building Permit Approval", "If substation enclosed", "https://www.dlt.go.th/"],
                ["Maintenance & Operation Plan", "Staffing, procedures", "N/A - Internal"],
            ]
        },
        "11. Water Supply Connection": {
            "Authority": "PWA / MWA",
            "FormLink": "https://www.mwa.co.th/ or https://www.pwa.co.th",
            "Items": [
                ["Water Connection Application", "Completed form", "https://www.mwa.co.th/ or https://www.pwa.co.th"],
                ["Site Location Map", "Showing connection point", "https://www.dlt.go.th/"],
                ["Water Requirement Calculation", "Daily/hourly demand", "N/A - Engineering"],
                ["Water Usage Estimate", "Peak and average consumption", "N/A - Engineering"],
                ["Building Plans", "Showing water routing", "N/A - Engineering"],
                ["Proof of Land Rights", "Title deed or lease", "https://www.dlt.go.th/"],
                ["Water Meter Specification", "Size and type", "N/A - Supplier"],
                ["Water Treatment Design", "If pre-treatment needed", "N/A - Engineering"],
                ["Pipe Routing Plan", "From connection to facility", "N/A - Engineering"],
                ["Environmental Assessment", "Water source & usage", "https://www.onep.go.th"],
                ["Company Registration Certificate", "Proof of legal entity", "https://www.dbd.go.th"],
            ]
        },
        "12. Wastewater / Drainage Connection": {
            "Authority": "Local Authority / Water Authority",
            "FormLink": "https://www.mwa.co.th/ or https://www.pwa.co.th",
            "Items": [
                ["Wastewater Discharge Application", "Formal request", "https://www.mwa.co.th/ or https://www.pwa.co.th"],
                ["Wastewater Treatment Design", "System specifications", "N/A - Engineering"],
                ["Discharge Quality Standards", "Compliance documentation", "https://www.onep.go.th"],
                ["Site Plan", "Showing discharge point", "https://www.dlt.go.th/"],
                ["Wastewater Characterization", "Expected pollutants, volumes", "N/A - Engineering"],
                ["Treatment System Specifications", "Equipment, capacity", "N/A - Engineering"],
                ["Drainage Design", "Pipes, sizing, routing", "N/A - Engineering"],
                ["Environmental Impact", "Discharge impact study", "https://www.onep.go.th"],
                ["Monitoring Plan", "Water quality testing schedule", "https://www.onep.go.th"],
                ["Building Plans", "Showing drainage system", "N/A - Engineering"],
                ["Proof of Land Rights", "Title deed or lease", "https://www.dlt.go.th/"],
            ]
        },
        "13. Fire Protection System Approval": {
            "Authority": "Local Building Control / Fire Authority",
            "FormLink": "https://www.dlt.go.th/",
            "Items": [
                ["Fire Safety Plan", "Comprehensive document", "https://www.dlt.go.th/"],
                ["Building Plans", "Fire exits, evacuation routes", "N/A - Engineering"],
                ["Fire Detection System Design", "Smoke/heat detectors", "N/A - Engineering"],
                ["Fire Suppression System Design", "Sprinkler/FM-200/Nitrogen", "N/A - Engineering"],
                ["Emergency Lighting Plan", "Exit signs, lighting", "N/A - Engineering"],
                ["Fire-Rated Materials", "Door, wall specifications", "N/A - Supplier"],
                ["Occupancy Load Calculation", "Number of occupants", "N/A - Engineering"],
                ["Evacuation Procedures", "Documented procedures", "N/A - Internal"],
                ["Staff Training Plan", "Fire safety training", "N/A - Internal"],
                ["Fire Alarm System Design", "Notification, monitoring", "N/A - Engineering"],
                ["Licensed Fire Safety Engineer", "Design certification", "N/A - Professional"],
            ]
        },
        "14. Generator & Fuel Storage": {
            "Authority": "Department of Energy Business (DOEB)",
            "FormLink": "https://www.doeb.go.th/",
            "Items": [
                ["Generator Specifications", "Capacity, fuel type, emissions", "https://www.doeb.go.th/"],
                ["Generator Installation Plan", "Location, foundation, access", "N/A - Engineering"],
                ["Fuel Storage Design", "Tank size, secondary containment", "N/A - Engineering"],
                ["Environmental Compliance", "Emissions testing plan", "https://www.onep.go.th"],
                ["Safety Plan", "Personnel protection, procedures", "N/A - Engineering"],
                ["Maintenance Schedule", "Regular testing & servicing", "N/A - Internal"],
                ["Site Plan", "Showing generator location", "https://www.dlt.go.th/"],
                ["Electrical Connection Design", "Integration with grid", "N/A - Engineering"],
                ["Noise & Vibration Assessment", "Mitigation measures", "N/A - Engineering"],
                ["Fuel Supply Agreement", "Regular fuel delivery", "N/A - Supplier"],
                ["As-Built Drawings", "After construction", "N/A - Engineering"],
                ["Emissions Testing Results", "Stack testing data", "https://www.onep.go.th"],
            ]
        },
        "15. Telecommunications / Fiber Connection": {
            "Authority": "NBTC (if licensed service)",
            "FormLink": "https://www.nbtc.go.th/",
            "Items": [
                ["Telecom Service Application", "Completed form", "https://www.nbtc.go.th/"],
                ["Network Architecture", "Fiber routing, equipment", "N/A - Engineering"],
                ["Service Area Map", "Coverage area", "N/A - Engineering"],
                ["Technical Specifications", "Bandwidth, redundancy", "N/A - Engineering"],
                ["License Application", "If providing public service", "https://www.nbtc.go.th/"],
                ["Security Plan", "Network security measures", "N/A - Engineering"],
                ["Backup Systems", "Redundant connections", "N/A - Engineering"],
                ["Service Level Agreement (SLA)", "Uptime guarantees", "N/A - Internal"],
                ["Carrier Agreements", "Fiber provider contracts", "N/A - Supplier"],
                ["Site Plans", "Showing fiber routes", "https://www.dlt.go.th/"],
                ["Environmental Assessment", "Cable routing impacts", "https://www.onep.go.th"],
            ]
        },
        "16. Electrical Inspection & Energization": {
            "Authority": "PEA / MEA",
            "FormLink": "https://www.pea.co.th/ or https://www.mea.or.th",
            "Items": [
                ["Final Electrical Design", "As-built drawings", "N/A - Engineering"],
                ["Electrical Inspection Request", "Inspection appointment", "https://www.pea.co.th/ or https://www.mea.or.th"],
                ["Equipment Certifications", "All electrical equipment", "N/A - Supplier"],
                ["Safety Compliance", "Grounding, breakers, protection", "N/A - Engineering"],
                ["Load Calculation Verification", "Final verification", "N/A - Engineering"],
                ["Meter Installation Plan", "Meter type and location", "https://www.pea.co.th/ or https://www.mea.or.th"],
                ["Substation Inspection", "Transformer, connections", "https://www.pea.co.th/ or https://www.mea.or.th"],
                ["Protective Relay Testing", "Verification of operation", "N/A - Engineering"],
                ["Energization Schedule", "Planned date", "https://www.pea.co.th/ or https://www.mea.or.th"],
                ["Operating Personnel Certification", "Staff training completion", "N/A - Internal"],
                ["Single-Line Diagram", "Final verified diagram", "N/A - Engineering"],
            ]
        },
        "17. Fire Safety Inspection": {
            "Authority": "Local Fire Authority",
            "FormLink": "https://www.dlt.go.th/",
            "Items": [
                ["Fire Safety Inspection Request", "Appointment scheduling", "https://www.dlt.go.th/"],
                ["Fire Suppression System", "Full system certification", "N/A - Engineering"],
                ["Fire Detection System", "Full system testing", "N/A - Engineering"],
                ["Emergency Lighting", "Operational verification", "N/A - Engineering"],
                ["Emergency Exits", "Accessibility, markings", "N/A - Engineering"],
                ["Fire-Rated Walls/Doors", "Installation verification", "N/A - Engineering"],
                ["Evacuation Procedures", "Posted and documented", "N/A - Internal"],
                ["Fire Extinguishers", "Properly mounted and charged", "N/A - Supplier"],
                ["Fire Hose Reels", "If equipped", "N/A - Supplier"],
                ["Stairwells & Routes", "Compliance verification", "N/A - Engineering"],
                ["Signage", "Fire safety signage complete", "N/A - Supplier"],
                ["Staff Training Records", "Fire safety training", "N/A - Internal"],
            ]
        },
        "18. Building Completion Inspection": {
            "Authority": "Local Building Control Authority",
            "FormLink": "https://www.dlt.go.th/",
            "Items": [
                ["Completion Inspection Request", "Application for inspection", "https://www.dlt.go.th/"],
                ["As-Built Drawings", "Final construction drawings", "N/A - Engineering"],
                ["Structural Certification", "Engineer sign-off", "N/A - Professional"],
                ["Material Certifications", "Fire-rated, non-toxic materials", "N/A - Supplier"],
                ["Safety Compliance", "All safety systems operational", "N/A - Engineering"],
                ["Accessibility Compliance", "Disabled access verified", "N/A - Engineering"],
                ["MEP Systems Testing", "HVAC, electrical, plumbing", "N/A - Engineering"],
                ["Quality Assurance Reports", "Inspection documentation", "N/A - Engineer"],
                ["Contractor Sign-off", "Defect list complete", "N/A - Contractor"],
                ["Environmental Compliance", "EIA compliance verification", "https://www.onep.go.th"],
                ["Utilities Connection", "All utilities connected", "https://www.pea.co.th/ or https://www.mea.or.th"],
            ]
        },
        "19. Occupancy Certificate": {
            "Authority": "Local Building Control Authority",
            "FormLink": "https://www.dlt.go.th/",
            "Items": [
                ["Building Completion Inspection", "Passed inspection", "https://www.dlt.go.th/"],
                ["Occupancy Permit Application", "Formal application", "https://www.dlt.go.th/"],
                ["Certificate of Occupancy Form", "Completed form", "https://www.dlt.go.th/"],
                ["Fire Safety Inspection", "Passed inspection", "https://www.dlt.go.th/"],
                ["Building Plans Compliance", "As-built vs. approved", "https://www.dlt.go.th/"],
                ["Utilities Operational", "Power, water, telecom active", "https://www.pea.co.th/ or https://www.mea.or.th"],
                ["Safety Systems Operational", "Fire, emergency systems", "N/A - Engineering"],
                ["Land Title/Lease", "Proof of legal occupancy right", "https://www.dlt.go.th/"],
                ["Business Registration", "Company legitimacy", "https://www.dbd.go.th"],
                ["Insurance Certificate", "Building insurance proof", "N/A - Insurance"],
                ["Environmental Clearance", "EIA compliance", "https://www.onep.go.th"],
            ]
        },
        "20. Commissioning & Utility Acceptance": {
            "Authority": "Utilities & Local Authorities",
            "FormLink": "https://www.pea.co.th/ / https://www.mea.or.th / https://www.mwa.co.th",
            "Items": [
                ["Electrical System Commissioning", "Full system testing", "https://www.pea.co.th/ or https://www.mea.or.th"],
                ["Water System Commissioning", "Pressure, quality testing", "https://www.mwa.co.th/ or https://www.pwa.co.th"],
                ["Wastewater System Commissioning", "Treatment system testing", "https://www.mwa.co.th/ or https://www.pwa.co.th"],
                ["Fire Protection Testing", "Full system operational test", "https://www.dlt.go.th/"],
                ["HVAC System Commissioning", "Temperature/humidity testing", "N/A - Engineering"],
                ["IT Infrastructure Commissioning", "Network, systems online", "N/A - Engineering"],
                ["Backup Power System Test", "Generator load test", "https://www.doeb.go.th/"],
                ["Utility Acceptance Documentation", "Utility sign-offs", "https://www.pea.co.th/ / https://www.mea.or.th"],
                ["Performance Certificates", "System performance verification", "N/A - Engineering"],
                ["Maintenance Documentation", "Service manuals, schedules", "N/A - Internal"],
                ["Staff Training Completion", "Operations team certified", "N/A - Internal"],
            ]
        },
        "21. Factory License": {
            "Authority": "Department of Industrial Works (DIW)",
            "FormLink": "https://www.diw.go.th/",
            "Items": [
                ["Factory License Application", "Completed form", "https://www.diw.go.th/"],
                ["Building Inspection Certificate", "From Local Authority", "https://www.dlt.go.th/"],
                ["Environmental Assessment", "EIA or notification", "https://www.onep.go.th"],
                ["Safety Equipment", "Fire, first aid, PPE", "N/A - Supplier"],
                ["Machinery Safety Certificates", "All equipment certified", "N/A - Supplier"],
                ["Hazard Assessment", "If manufacturing on-site", "N/A - Engineering"],
                ["Worker Safety Plan", "Training, procedures", "N/A - Internal"],
                ["Waste Management Plan", "Industrial waste handling", "https://www.onep.go.th"],
                ["Site Plan", "Showing factory area", "https://www.dlt.go.th/"],
                ["Manager Appointment", "Factory manager information", "https://www.diw.go.th/"],
                ["Electrical Safety Inspection", "DIW electrical compliance", "https://www.diw.go.th/"],
                ["Occupancy Permit", "Building completion", "https://www.dlt.go.th/"],
            ]
        },
        "22. PDPA Compliance": {
            "Authority": "Personal Data Protection Committee",
            "FormLink": "https://www.pdpa.go.th/",
            "Items": [
                ["Data Protection Policy", "Written policy document", "https://www.pdpa.go.th/"],
                ["Privacy Notice", "For data subjects", "https://www.pdpa.go.th/"],
                ["Data Processing Agreement", "With 3rd parties", "https://www.pdpa.go.th/"],
                ["Data Breach Response Plan", "Incident procedures", "https://www.pdpa.go.th/"],
                ["Staff Training Records", "PDPA training completion", "https://www.pdpa.go.th/"],
                ["Consent Documentation", "Personal data consent forms", "https://www.pdpa.go.th/"],
                ["Data Retention Schedule", "Data destruction dates", "N/A - Internal"],
                ["Data Subject Rights Procedures", "Access request procedures", "https://www.pdpa.go.th/"],
                ["Vendor Assessment", "3rd party data processor", "N/A - Internal"],
                ["Data Mapping", "What data, where stored", "N/A - Internal"],
                ["Security Measures", "Encryption, access controls", "N/A - Engineering"],
            ]
        },
        "23. Cybersecurity Compliance": {
            "Authority": "National Cyber Security Agency",
            "FormLink": "https://www.ncsa.go.th/",
            "Items": [
                ["Critical Infrastructure Registration", "If designated", "https://www.ncsa.go.th/"],
                ["Cybersecurity Assessment", "Professional assessment", "https://www.ncsa.go.th/"],
                ["Security Controls Documentation", "Technical controls", "https://www.ncsa.go.th/"],
                ["Incident Response Plan", "Cyber incident procedures", "https://www.ncsa.go.th/"],
                ["Staff Training Records", "Cybersecurity training", "https://www.ncsa.go.th/"],
                ["Penetration Testing Results", "Annual testing results", "N/A - Consultant"],
                ["Vulnerability Assessment Report", "Regular scans", "N/A - Consultant"],
                ["Backup & Recovery Plan", "Data continuity procedures", "https://www.ncsa.go.th/"],
                ["Access Control Documentation", "User authentication system", "N/A - Engineering"],
                ["Security Monitoring", "SIEM or logging system", "N/A - Engineering"],
                ["Disaster Recovery Plan", "Business continuity", "https://www.ncsa.go.th/"],
                ["Compliance Certification", "SOC 2, ISO 27001", "N/A - Certifier"],
            ]
        },
    }

    # Create master index sheet
    ws_index = wb.create_sheet("00_INDEX", 0)
    ws_index['A1'] = "PERMIT CHECKLISTS - WHAT TO PROVIDE TO EACH AUTHORITY"
    ws_index['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_index['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws_index.merge_cells('A1:D1')

    ws_index['A3'] = "Click on each permit tab below to see detailed checklist"
    ws_index.merge_cells('A3:D3')

    ws_index['A5'] = "Permit Number"
    ws_index['B5'] = "Permit Name"
    ws_index['C5'] = "Authority"
    ws_index['D5'] = "Form Link"

    for col in ['A', 'B', 'C', 'D']:
        ws_index[f'{col}5'].font = Font(bold=True, color="FFFFFF")
        ws_index[f'{col}5'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Add permits to index
    row = 6
    for permit_name, permit_info in checklists.items():
        parts = permit_name.split(". ", 1)
        ws_index[f'A{row}'] = parts[0]
        ws_index[f'B{row}'] = parts[1] if len(parts) > 1 else ""
        ws_index[f'C{row}'] = permit_info['Authority']

        # Add hyperlink for form
        cell = ws_index[f'D{row}']
        cell.value = "Form Link"
        cell.hyperlink = permit_info['FormLink']
        cell.font = Font(color="0563C1", underline="single")

        row += 1

    ws_index.column_dimensions['A'].width = 5
    ws_index.column_dimensions['B'].width = 40
    ws_index.column_dimensions['C'].width = 40
    ws_index.column_dimensions['D'].width = 20

    # Create checklist sheet for each permit
    for permit_name, permit_info in checklists.items():
        sheet_name = permit_name.replace(" ", "_").replace("(", "").replace(")", "").replace("/", "-")[:31]
        ws = wb.create_sheet(sheet_name)

        # Header
        ws['A1'] = permit_name
        ws['A1'].font = Font(size=12, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        ws.merge_cells('A1:F1')

        ws['A2'] = f"Authority: {permit_info['Authority']}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:F2')

        ws['A3'] = f"Main Form Link: {permit_info['FormLink']}"
        ws['A3'].font = Font(size=9, italic=True, color="0563C1")
        ws.merge_cells('A3:F3')

        # Column headers
        headers = ["Item #", "Document / Item Required", "Format / Requirement", "Form Link / Resources", "Status", "Date Submitted"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(5, col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[5].height = 25

        # Add checklist items
        for item_idx, item in enumerate(permit_info['Items'], 1):
            row_num = 5 + item_idx
            ws.cell(row_num, 1).value = item_idx
            ws.cell(row_num, 2).value = item[0]
            ws.cell(row_num, 3).value = item[1]

            # Add hyperlink for form
            form_cell = ws.cell(row_num, 4)
            if item[2] and item[2] != "N/A - Internal" and item[2] != "N/A - Engineering" and item[2] != "N/A - Professional" and item[2] != "N/A - Supplier" and item[2] != "N/A - Bank Document" and item[2] != "N/A - Auditor" and item[2] != "N/A - Consultant" and item[2] != "N/A - Contractor" and item[2] != "N/A - Insurance" and item[2] != "N/A - Certifier" and item[2] != "N/A - Government ID" and item[2] != "N/A - Company Document" and item[2] != "N/A - Legal Document" and not item[2].startswith("N/A"):
                form_cell.value = "Online Form"
                form_cell.hyperlink = item[2]
                form_cell.font = Font(color="0563C1", underline="single", size=9)
            else:
                form_cell.value = item[2]
                form_cell.font = Font(size=9, italic=True, color="666666")

            ws.cell(row_num, 5).value = "Not Done"
            ws.cell(row_num, 6).value = ""

            # Format status column
            status_cell = ws.cell(row_num, 5)
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Align other cells
            for col in [1, 2, 3, 4, 6]:
                cell = ws.cell(row_num, col)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.font = Font(size=9)

        # Format columns
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14

    # Save workbook
    wb.save('PERMIT_CHECKLISTS_ALL_23.xlsx')
    print("SUCCESS! Created: PERMIT_CHECKLISTS_ALL_23.xlsx")
    print(f"\nSheets Created: {len(wb.sheetnames)}")
    print("[OK] INDEX sheet with form links navigation")
    print("[OK] 23 Individual permit sheets with:")
    print("     - Form links for documents")
    print("     - Hyperlinked resources to government portals")
    print("     - Status tracking columns")
    print("\nFeatures:")
    print("* Click on authority form links in INDEX sheet")
    print("* Each permit has main Form Link at top")
    print("* Individual items show specific form links (blue=clickable)")
    print("* N/A items show type (Engineering, Internal, etc)")

create_permit_checklists_with_forms()
