from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def create_permit_manager_tracker():
    """Create comprehensive permit manager tracker with all 23 permits"""
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Define colors by phase
    phase_colors = {
        'Development': 'D9E1F2',
        'Design/Construction': 'E2EFDA',
        'Construction': 'FCE4D6',
        'Commissioning': 'FFF2CC',
        'Operation': 'F4B084'
    }

    # All 23 permits data
    permits_data = [
        [1, "Company Registration", "Department of Business Development", "Development", 30, "Not Started", "", "", ""],
        [2, "BOI Promotion Certificate (if applicable)", "Thailand Board of Investment", "Development", 120, "Not Started", "", "", ""],
        [3, "Foreign Business License (if applicable)", "Department of Business Development", "Development", 60, "Not Started", "", "", ""],
        [4, "Land Use & Zoning Confirmation", "Dept of Public Works / Local Authority", "Development", 30, "Not Started", "", "", ""],
        [5, "Environmental Impact Assessment (EIA) Approval", "Office of Natural Resources & Environmental Policy", "Development", 180, "Not Started", "", "", ""],
        [6, "Building Construction Permit", "Local Building Control Authority", "Design/Construction", 60, "Not Started", "", "", ""],
        [7, "Road Access / Highway Connection Approval", "Local Authority / Highway Authority", "Construction", 45, "Not Started", "", "", ""],
        [8, "Grid Connection Approval", "PEA or MEA", "Construction", 60, "Not Started", "", "", ""],
        [9, "Power Availability Confirmation", "Energy Regulatory Commission", "Development", 30, "Not Started", "", "", ""],
        [10, "High-Voltage Substation Approval", "PEA / MEA", "Construction", 90, "Not Started", "", "", ""],
        [11, "Water Supply Connection Approval", "PWA or MWA", "Construction", 45, "Not Started", "", "", ""],
        [12, "Wastewater / Drainage Connection Approval", "Local Authority / Water Authority", "Construction", 45, "Not Started", "", "", ""],
        [13, "Fire Protection System Approval", "Local Building Control / Fire Authority", "Construction", 30, "Not Started", "", "", ""],
        [14, "Generator & Fuel Storage Approval", "Department of Energy Business", "Construction", 60, "Not Started", "", "", ""],
        [15, "Telecommunications / Fiber Connection Approval", "National Broadcasting & Telecommunications Commission", "Construction", 45, "Not Started", "", "", ""],
        [16, "Electrical Installation Inspection & Energization", "PEA / MEA", "Commissioning", 14, "Not Started", "", "", ""],
        [17, "Fire Safety Inspection", "Local Fire Authority", "Commissioning", 7, "Not Started", "", "", ""],
        [18, "Building Completion Inspection", "Local Building Control Authority", "Commissioning", 7, "Not Started", "", "", ""],
        [19, "Occupancy Certificate / Building Use Permit", "Local Building Control Authority", "Commissioning", 14, "Not Started", "", "", ""],
        [20, "Commissioning & Utility Acceptance", "Utilities and Local Authorities", "Commissioning", 21, "Not Started", "", "", ""],
        [21, "Factory License (if applicable)", "Department of Industrial Works", "Operation", 30, "Not Started", "", "", ""],
        [22, "PDPA Compliance", "Personal Data Protection Committee", "Operation", 30, "Not Started", "", "", ""],
        [23, "Cybersecurity Compliance (Critical Infrastructure)", "National Cyber Security Agency", "Operation", 30, "Not Started", "", "", ""],
    ]

    # ==================== SHEET 1: DASHBOARD ====================
    ws_dash = wb.create_sheet("Dashboard", 0)

    ws_dash['A1'] = "PERMIT MANAGER'S DASHBOARD - CHN1A PROJECT"
    ws_dash['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_dash['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws_dash.merge_cells('A1:F1')
    ws_dash.row_dimensions[1].height = 25

    ws_dash['A2'] = f"Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_dash['A2'].font = Font(italic=True, size=9)

    # Phase summary
    ws_dash['A4'] = "PERMITS BY PHASE"
    ws_dash['A4'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_dash['A4'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_dash.merge_cells('A4:F4')

    phases = [
        ["Development", 5, "=COUNTIF('Master List'!D:D,\"Done\")", "% Complete"],
        ["Design/Construction", 2, "=COUNTIF('Master List'!D:D,\"Done\")", "% Complete"],
        ["Construction", 8, "=COUNTIF('Master List'!D:D,\"Done\")", "% Complete"],
        ["Commissioning", 5, "=COUNTIF('Master List'!D:D,\"Done\")", "% Complete"],
        ["Operation", 3, "=COUNTIF('Master List'!D:D,\"Done\")", "% Complete"],
    ]

    row = 5
    for phase_info in phases:
        ws_dash.cell(row, 1).value = phase_info[0]
        ws_dash.cell(row, 1).font = Font(bold=True)
        ws_dash.cell(row, 2).value = f"Total: {phase_info[1]}"
        ws_dash.cell(row, 3).value = "Done: "
        ws_dash.cell(row, 4).value = "In Progress: "
        ws_dash.cell(row, 5).value = "Pending: "
        row += 1

    # Critical path
    row += 2
    ws_dash.cell(row, 1).value = "CRITICAL PATH ALERTS"
    ws_dash.cell(row, 1).font = Font(bold=True, size=12, color="FFFFFF")
    ws_dash.cell(row, 1).fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws_dash.merge_cells(f'A{row}:F{row}')

    row += 1
    alerts = [
        ["CRITICAL", "BOI Certificate", "Dependency gate for many permits", "In Progress"],
        ["CRITICAL", "EIA Approval", "Blocks Building Permit", "Not Started"],
        ["URGENT", "Power Confirmation", "Grid connection depends on this", "Not Started"],
        ["HIGH", "Building Permit", "Gates all construction permits", "Not Started"],
    ]

    for alert in alerts:
        for col, val in enumerate(alert, 1):
            ws_dash.cell(row, col).value = val
            if col == 1 and alert[0] == "CRITICAL":
                ws_dash.cell(row, col).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws_dash.cell(row, col).font = Font(color="FFFFFF", bold=True)
        row += 1

    ws_dash.column_dimensions['A'].width = 20
    ws_dash.column_dimensions['B'].width = 25
    ws_dash.column_dimensions['C'].width = 25
    ws_dash.column_dimensions['D'].width = 15
    ws_dash.column_dimensions['E'].width = 15
    ws_dash.column_dimensions['F'].width = 15

    # ==================== SHEET 2: MASTER LIST ====================
    ws_master = wb.create_sheet("Master List", 1)

    ws_master['A1'] = "MASTER PERMIT LIST - ALL 23 PERMITS"
    ws_master['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_master['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws_master.merge_cells('A1:J1')

    headers = ["#", "Permit Name", "Authority", "Phase", "Lead Time (Days)", "Status", "Planned Start", "Deadline", "Actual Completion", "Owner", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws_master.cell(3, col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_master.row_dimensions[3].height = 25

    # Add permits
    for row_idx, permit in enumerate(permits_data, 4):
        for col_idx, value in enumerate(permit, 1):
            cell = ws_master.cell(row_idx, col_idx)
            cell.value = value

            # Phase color coding
            phase = permit[3]
            if phase in phase_colors:
                cell.fill = PatternFill(start_color=phase_colors[phase], end_color=phase_colors[phase], fill_type="solid")

            # Status coloring
            if col_idx == 6:  # Status column
                if value == "Done":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(bold=True)
                elif value == "In Progress":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(bold=True)
                elif value == "Not Started":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Format columns
    ws_master.column_dimensions['A'].width = 3
    ws_master.column_dimensions['B'].width = 40
    ws_master.column_dimensions['C'].width = 30
    ws_master.column_dimensions['D'].width = 18
    ws_master.column_dimensions['E'].width = 15
    ws_master.column_dimensions['F'].width = 14
    ws_master.column_dimensions['G'].width = 12
    ws_master.column_dimensions['H'].width = 12
    ws_master.column_dimensions['I'].width = 14
    ws_master.column_dimensions['J'].width = 12
    ws_master.column_dimensions['K'].width = 20

    # ==================== SHEET 3: AUTHORITY CONTACTS ====================
    ws_auth = wb.create_sheet("Authority Contacts", 2)

    ws_auth['A1'] = "AUTHORITY CONTACT INFORMATION"
    ws_auth['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_auth['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws_auth.merge_cells('A1:G1')

    auth_headers = ["Authority Name", "Phone", "Email", "Office Address", "Coordinator Name", "Office Hours", "Website"]
    for col, header in enumerate(auth_headers, 1):
        cell = ws_auth.cell(3, col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    authorities = [
        ["Department of Business Development (DBD)", "+66-2-547-3000", "info@dbd.go.th", "Bangkok", "TBD", "08:30-16:30", "www.dbd.go.th"],
        ["Thailand Board of Investment (BOI)", "+66-2-553-8111", "inquiry@boi.go.th", "Bangkok", "TBD", "08:30-16:30", "www.boi.go.th"],
        ["Office of Natural Resources & Environmental Policy (ONEP)", "+66-2-535-8800", "inquiry@onep.go.th", "Bangkok", "TBD", "08:30-16:30", "www.onep.go.th"],
        ["Local Building Control Authority", "TBD", "TBD", "Local", "TBD", "TBD", "TBD"],
        ["Provincial Electricity Authority (PEA)", "+66-2-399-9999", "TBD", "Regional", "TBD", "TBD", "www.pea.co.th"],
        ["Metropolitan Electricity Authority (MEA)", "+66-2-246-3888", "TBD", "Bangkok", "TBD", "08:30-16:30", "www.mea.or.th"],
        ["Energy Regulatory Commission (ERC)", "+66-2-553-8511", "TBD", "Bangkok", "TBD", "08:30-16:30", "www.erc.or.th"],
        ["Provincial Waterworks Authority (PWA)", "+66-2-503-9000", "TBD", "Regional", "TBD", "08:30-16:30", "www.pwa.co.th"],
        ["National Broadcasting & Telecommunications Commission (NBTC)", "+66-2-575-9000", "TBD", "Bangkok", "TBD", "08:30-16:30", "www.nbtc.go.th"],
        ["Department of Energy Business (DOEB)", "+66-2-553-6500", "TBD", "Bangkok", "TBD", "08:30-16:30", "www.doeb.go.th"],
        ["Department of Industrial Works (DIW)", "+66-2-547-1310", "TBD", "Bangkok", "TBD", "08:30-16:30", "www.diw.go.th"],
        ["National Cyber Security Agency", "+66-2-141-2999", "TBD", "Bangkok", "TBD", "08:30-16:30", "TBD"],
        ["Personal Data Protection Committee", "TBD", "TBD", "Bangkok", "TBD", "TBD", "TBD"],
    ]

    for row_idx, auth in enumerate(authorities, 4):
        for col_idx, value in enumerate(auth, 1):
            cell = ws_auth.cell(row_idx, col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws_auth.column_dimensions['A'].width = 35
    ws_auth.column_dimensions['B'].width = 15
    ws_auth.column_dimensions['C'].width = 20
    ws_auth.column_dimensions['D'].width = 20
    ws_auth.column_dimensions['E'].width = 15
    ws_auth.column_dimensions['F'].width = 15
    ws_auth.column_dimensions['G'].width = 20

    # ==================== SHEET 4: ACTION ITEMS ====================
    ws_actions = wb.create_sheet("Action Items", 3)

    ws_actions['A1'] = "ACTION ITEMS & TASKS - THIS WEEK"
    ws_actions['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_actions['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws_actions.merge_cells('A1:G1')

    action_headers = ["Task", "Permit #", "Owner", "Due Date", "Priority", "Status", "Notes"]
    for col, header in enumerate(action_headers, 1):
        cell = ws_actions.cell(3, col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    actions = [
        ["Verify ERC Power Confirmation Letter", "9", "TBD", datetime.now().strftime('%Y-%m-%d'), "CRITICAL", "In Progress", "Prerequisite for BOI"],
        ["Complete BOI Application Documents", "2", "TBD", (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d'), "CRITICAL", "In Progress", ""],
        ["Submit EIA Screening", "5", "TBD", (datetime.now() + timedelta(days=3)).strftime('%Y-%m-%d'), "CRITICAL", "Not Started", ""],
        ["Prepare FBL Documents", "3", "TBD", (datetime.now() + timedelta(days=14)).strftime('%Y-%m-%d'), "HIGH", "Not Started", "After BOI approval"],
        ["Contact DBD for Company Registration", "1", "TBD", (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d'), "HIGH", "Not Started", ""],
    ]

    for row_idx, action in enumerate(actions, 4):
        for col_idx, value in enumerate(action, 1):
            cell = ws_actions.cell(row_idx, col_idx)
            cell.value = value
            if col_idx == 5 and value == "CRITICAL":
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            elif col_idx == 5 and value == "HIGH":
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    ws_actions.column_dimensions['A'].width = 35
    ws_actions.column_dimensions['B'].width = 10
    ws_actions.column_dimensions['C'].width = 12
    ws_actions.column_dimensions['D'].width = 12
    ws_actions.column_dimensions['E'].width = 12
    ws_actions.column_dimensions['F'].width = 12
    ws_actions.column_dimensions['G'].width = 25

    # ==================== SHEET 5: RISKS & ISSUES ====================
    ws_risks = wb.create_sheet("Risks & Issues", 4)

    ws_risks['A1'] = "RISKS & ISSUES BLOCKING PERMITS"
    ws_risks['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws_risks['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws_risks.merge_cells('A1:H1')

    risk_headers = ["Permit #", "Issue/Blocker", "Severity", "Owner", "Mitigation Plan", "Status", "Target Resolution Date", "Notes"]
    for col, header in enumerate(risk_headers, 1):
        cell = ws_risks.cell(3, col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    risks = [
        ["9", "ERC power confirmation delayed", "CRITICAL", "TBD", "Escalate to Thailand FastPass", "In Progress", (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d'), "Gate for BOI application"],
        ["5", "EIA study scope unclear", "HIGH", "TBD", "Clarify with ONEP", "Not Started", (datetime.now() + timedelta(days=14)).strftime('%Y-%m-%d'), ""],
        ["2", "BOI capital requirements", "MEDIUM", "TBD", "Verify THB 2B+ threshold met", "Not Started", (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d'), ""],
    ]

    for row_idx, risk in enumerate(risks, 4):
        for col_idx, value in enumerate(risk, 1):
            cell = ws_risks.cell(row_idx, col_idx)
            cell.value = value
            if col_idx == 3 and value == "CRITICAL":
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            elif col_idx == 3 and value == "HIGH":
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    ws_risks.column_dimensions['A'].width = 10
    ws_risks.column_dimensions['B'].width = 25
    ws_risks.column_dimensions['C'].width = 12
    ws_risks.column_dimensions['D'].width = 12
    ws_risks.column_dimensions['E'].width = 25
    ws_risks.column_dimensions['F'].width = 12
    ws_risks.column_dimensions['G'].width = 18
    ws_risks.column_dimensions['H'].width = 25

    # Save workbook
    wb.save('PERMIT_MANAGER_TRACKER.xlsx')
    print("SUCCESS! Created: PERMIT_MANAGER_TRACKER.xlsx")
    print("\nSheets Created:")
    for idx, sheet in enumerate(wb.sheetnames, 1):
        print(f"  {idx}. {sheet}")

create_permit_manager_tracker()
